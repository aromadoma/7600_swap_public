import click
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from config_generator.generator.gentools import format_data_from_table
from config_generator.fstools import create_preparing_dir, get_homedir_path, \
    get_billing_path


def create_preparation_table(workbook_name,
                             vlan_keys=None,
                             vlan_data=None,
                             vrf_keys=None,
                             vrf_data=None,
                             xc_keys=None,
                             xc_data=None,
                             phys_int_keys=None,
                             phys_int_data=None,
                             unique_vrf_keys=None,
                             unique_vrf_data=None,
                             unique_vlan_keys=None,
                             unique_vlan_data=None,
                             neighbor_keys=None,
                             neighbor_data=None
                             ):
    workbook = Workbook()  # Создаем таблицу
    if vlan_keys and vlan_data:
        create_sheet(workbook, 'Список вланов', vlan_keys, vlan_data)
    if vrf_keys and vrf_data:
        create_sheet(workbook, 'Список VRF', vrf_keys, vrf_data)
    if xc_keys and xc_data:
        create_sheet(workbook, 'Список XC', xc_keys, xc_data)
    if phys_int_keys and phys_int_data:
        create_sheet(workbook, 'Список интерфейсов', phys_int_keys, phys_int_data)
    if unique_vlan_keys and unique_vlan_data:
        create_sheet(workbook, 'Общий список VLAN', unique_vlan_keys, unique_vlan_data)
    if unique_vrf_keys and unique_vrf_data:
        create_sheet(workbook, 'Общий список VRF', unique_vrf_keys, unique_vrf_data)
    if neighbor_keys and neighbor_data:
        create_sheet(workbook, 'Общий список соседей', neighbor_keys, neighbor_data)

    # Сохраняем таблицу:
    workbook_path = create_preparing_dir(get_homedir_path())
    while True:
        # В случае уже открытого excel-файла ожидаем закрытия:
        try:
            workbook.save(f'{workbook_path}/{workbook_name}')
            break
        except PermissionError:
            click.echo('\u001b[31mВНИМАНИЕ!\u001b[0m '
                       'Не удается сохранить excel-файл, вероятно, он открыт.\nЗакройте '
                       'файл и нажмите ENTER', nl=False)
            input()


def create_sheet(wb, name, keys, data):
    """

    :param wb: workbook obj
    :param name: str, name of sheet
    :param keys: list, keys for dictionaries in data variable
    :param data: list of dictionaries with data
    :return:
    """
    sheet = wb.create_sheet(name)
    sheet.title = name

    # Заголовки:
    for i in range(len(keys)):
        sheet.cell(row=1, column=i + 1, value=keys[i])
        sheet[1][i].font = Font(bold=True)
        sheet[1][i].alignment = Alignment(horizontal='center', vertical='center')
        sheet[1][i].fill = PatternFill("solid", fgColor='00339966')

    # Данные:
    row_number = 2
    for element in data:
        for column_number in range(1, len(keys) + 1):
            value = element.get(keys[column_number - 1])
            if isinstance(value, list):  # Объединяем, если список
                sheet.cell(row=row_number, column=column_number).value = ', '.join(value)
            else:
                sheet.cell(row=row_number, column=column_number).value = value

    # Подсветка строк с флагом ALARM:
        if element.get('alarm'):
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor='00FF0000')
        row_number += 1

    # Фильтр на все столбцы:
    last_column = sheet[1][sheet.max_column - 1].column_letter
    last_row = row_number - 1
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    # Замораживаем заголовки:
    sheet.freeze_panes = 'A2'

    return None


def get_billing_data(hostname):
    billing_path = get_billing_path(hostname)
    workbook = load_workbook(filename=billing_path)
    sheet = workbook.active

    vlan_list = []
    for row in range(5, sheet.max_row):
        if sheet['A{}'.format(row)].value.find('Vlan') != -1:
            vlan_list.append(sheet['A{}'.format(row)].value.replace('Vlan', ''))
        row += 1

    workbook.close()
    return vlan_list


def load_data_from_table(sheets, keys_list):
    data = []
    for sheet, keys in zip(sheets, keys_list):
        data.append(format_data_from_table(sheet.title, get_data_from_sheet(sheet, keys)))
    return data


def get_data_from_sheet(sheet, keys):
    element = {}
    data_list = []
    for row_num in range(2, sheet.max_row + 1):
        for column_num in range(1, len(keys) + 1):
            element[keys[column_num - 1]] = \
                sheet.cell(row=row_num, column=column_num).value
        data_list.append(element)
        element = {}
    return data_list
